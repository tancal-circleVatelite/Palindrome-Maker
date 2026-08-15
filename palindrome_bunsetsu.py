
from __future__ import annotations

"""
palindrome_bunsetsu.py
======================

日本語の文節集合（CSV）から、回文候補を探索するスクリプトです。

この実装で重視していること
----------------------------
1. 回文生成アルゴリズムの学習用に、処理の意味がわかる丁寧なコメントを付ける
2. コマンドラインから簡単に実行できる
3. 生成結果を Excel ブック (.xlsx) に保存できる
4. 候補一覧に簡易自然度スコアを記録できる
5. サマリーに総探索ノード数を記録できる
6. 実行中に、直近で生成した回文候補を標準出力に上書き表示できる

今回の出力仕様
--------------
- 生成結果は .xlsx 形式で保存する
- 入力した文節集合 CSV と同じフォルダへ保存する
- ファイル名は「シード文節_出力日.xlsx」とする
- Excel ブックは次の 2 シート構成
  - 「一覧」: 生成した回文候補一覧
  - 「サマリー」: 処理時間、候補数、総探索ノード数、実行時引数など

表示仕様
--------
- 探索中、回文候補が 1 件見つかるたびに、標準出力へ「最新の候補」を 1 行で上書き表示する
- 改行しないで更新していくため、ターミナル上では 1 行だけが変化して見える
- 最後に候補一覧を表示する前に改行して整形する

注意
----
- 入力 CSV の必須列は text, reading です。
- btype 列は任意です。
- 出力日の形式は YYYYMMDD を採用しています。
"""

from dataclasses import dataclass, field
from typing import Callable, List, Optional, Sequence, Set, Tuple
import csv
import argparse
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# =============================================================================
# データ構造
# =============================================================================

@dataclass(frozen=True)
class Bunsetsu:
    """
    文節エントリを表すデータクラス。

    Parameters
    ----------
    text : str
        文節の表記（例: 「夜」「起きぬ」「タヌキ」）
    reading : str
        文節の読み（例: 「よる」「おきぬ」「たぬき」）
    btype : str
        文節タイプ（例: noun / verb / aux など）
        現時点では強い文法制約には使っていませんが、
        将来の接続条件追加に備えて保持してあります。
    """
    text: str
    reading: str
    btype: str = ""


@dataclass(frozen=True)
class State:
    """
    探索中の状態を表すデータクラス。

    状態は <L, H, R> で管理する。

    L : 左不足文字列
        現在の読み列の左側に、回文条件を満たすためにまだ必要な文字列
    H : 現在の読み列
        すでに連結された文節全体の読み
    R : 右不足文字列
        現在の読み列の右側に、回文条件を満たすためにまだ必要な文字列
    parts : Tuple[Bunsetsu, ...]
        いままでに連結した文節列本体
    """
    L: str
    H: str
    R: str
    parts: Tuple[Bunsetsu, ...] = field(default_factory=tuple)

    def is_final(self) -> bool:
        """左右の不足文字列がともに空なら、読みとしては回文条件を満たしている。"""
        return not self.L and not self.R

    def reading(self) -> str:
        """現在の読み列を返す。"""
        return self.H

    def text(self) -> str:
        """現在の文節列の表記を連結して返す。"""
        return "".join(p.text for p in self.parts)


# =============================================================================
# CSV 読み込み
# =============================================================================

def load_bunsetsu_csv(csv_path: str, encoding: str = "utf-8-sig") -> List[Bunsetsu]:
    """
    文節集合 CSV を読み込んで、Bunsetsu のリストを返す。

    想定ヘッダ
    ----------
    必須:
      - text
      - reading
    任意:
      - btype
    """
    out: List[Bunsetsu] = []

    with open(csv_path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)

        if reader.fieldnames is None:
            raise ValueError(
                "CSVヘッダが見つかりません。1行目に text,reading,btype などのヘッダを入れてください。"
            )

        normalized = {name.strip().lower(): name for name in reader.fieldnames}
        if "text" not in normalized or "reading" not in normalized:
            raise ValueError(
                f"CSVヘッダに text と reading が必要です。現在のヘッダ: {reader.fieldnames}"
            )

        text_key = normalized["text"]
        reading_key = normalized["reading"]
        btype_key = normalized.get("btype")

        for i, row in enumerate(reader, start=2):
            text = (row.get(text_key) or "").strip()
            reading = (row.get(reading_key) or "").strip()
            btype = (row.get(btype_key) or "").strip() if btype_key else ""

            if not text and not reading:
                continue
            if not text or not reading:
                raise ValueError(f"{i}行目: text または reading が空です。 row={row}")

            out.append(Bunsetsu(text=text, reading=reading, btype=btype))

    if not out:
        raise ValueError("CSVから文節が1件も読み込めませんでした。")

    return out


# =============================================================================
# 初期状態生成（折返し固定法ベース）
# =============================================================================

def _initial_state_for_axis(seed: Bunsetsu, axis2: int) -> Optional[State]:
    """
    シード文節 seed.reading の中に折返し軸を 1 本仮定し、そこから初期状態を 1 つ作る。
    """
    s = seed.reading
    m = len(s)
    left_out: List[str] = []
    right_out: List[str] = []

    for i, ch in enumerate(s):
        c2 = 2 * i + 1
        mc2 = 2 * axis2 - c2

        if mc2 < 1:
            left_out.append(ch)
        elif mc2 > 2 * m - 1:
            right_out.append(ch)
        elif mc2 % 2 == 1:
            j = (mc2 - 1) // 2
            if s[j] != ch:
                return None
        else:
            return None

    L = "".join(reversed(left_out))
    R = "".join(reversed(right_out))
    if L and R:
        return None

    return State(L=L, H=s, R=R, parts=(seed,))



def initial_states(seed: Bunsetsu) -> List[State]:
    """シード文節から取り得る全ての初期状態を列挙する。"""
    s = seed.reading
    m = len(s)
    out: List[State] = []
    seen: Set[Tuple[str, str, str]] = set()

    for axis2 in range(0, 2 * m + 1):
        st = _initial_state_for_axis(seed, axis2)
        if st is not None:
            key = (st.L, st.H, st.R)
            if key not in seen:
                seen.add(key)
                out.append(st)

    return out


# =============================================================================
# 基本オペレータ（左右に文節を足す）
# =============================================================================

def extend_left(state: State, w: Bunsetsu) -> Optional[State]:
    """現在状態の左側に文節 w を追加できるか判定し、できれば新しい状態を返す。"""
    L = state.L
    if not L:
        return None

    wr = w.reading

    if len(L) >= len(wr) and L.endswith(wr):
        newL = L[:-len(wr)] if len(wr) else L
        return State(L=newL, H=wr + state.H, R="", parts=(w,) + state.parts)

    if wr.endswith(L):
        prefix_part = wr[:-len(L)] if len(L) else wr
        newR = prefix_part[::-1]
        return State(L="", H=wr + state.H, R=newR, parts=(w,) + state.parts)

    return None



def extend_right(state: State, w: Bunsetsu) -> Optional[State]:
    """現在状態の右側に文節 w を追加できるか判定し、できれば新しい状態を返す。"""
    R = state.R
    if not R:
        return None

    wr = w.reading

    if len(R) >= len(wr) and R.startswith(wr):
        newR = R[len(wr):]
        return State(L="", H=state.H + wr, R=newR, parts=state.parts + (w,))

    if wr.startswith(R):
        suffix_part = wr[len(R):]
        newL = suffix_part[::-1]
        return State(L=newL, H=state.H + wr, R="", parts=state.parts + (w,))

    return None


# =============================================================================
# 文法チェック用フック
# =============================================================================
CanConnect = Callable[[Optional[Bunsetsu], Bunsetsu, str], bool]
FinalOK = Callable[[Tuple[Bunsetsu, ...]], bool]
ProgressCallback = Callable[[int, int], None]
ResultCallback = Callable[[State], None]


def default_can_connect(neighbor: Optional[Bunsetsu], new_item: Bunsetsu, side: str) -> bool:
    return True


def default_final_ok(parts: Tuple[Bunsetsu, ...]) -> bool:
    return True


# =============================================================================
# 自然度スコア
# =============================================================================

def compute_naturality_score(parts: Tuple[Bunsetsu, ...]) -> int:
    """
    文節列に対して、簡易的な自然度スコアを計算する。

    このスコアは厳密な意味理解ではなく、
    「文として比較的自然そうか」をざっくり数値化するためのヒューリスティックである。
    """
    if not parts:
        return 0

    score = 0
    btypes = [p.btype for p in parts]

    last = btypes[-1]
    if last in {"verb", "aux"}:
        score += 3
    elif last == "adj":
        score += 2
    elif last == "noun":
        score -= 1

    distinct_types = {b for b in btypes if b}
    score += min(len(distinct_types), 3)

    if "aux" in btypes:
        score += 1
    if "particle" in btypes:
        score += 2

    for a, b in zip(btypes, btypes[1:]):
        if a and a == b:
            score -= 1

    if len(distinct_types) == 1 and len(parts) >= 2:
        score -= 2

    if sum(len(p.reading) for p in parts) <= len(parts) * 2:
        score -= 1

    return score


# =============================================================================
# メイン探索関数
# =============================================================================

def generate_palindrome_candidates(
    seed: Bunsetsu,
    db: Sequence[Bunsetsu],
    target_bunsetsu: int,
    can_connect: CanConnect = default_can_connect,
    final_ok: FinalOK = default_final_ok,
    allow_duplicate_parts: bool = True,
    progress_callback: Optional[ProgressCallback] = None,
    result_callback: Optional[ResultCallback] = None,
    cancel_callback: Optional[Callable[[], bool]] = None,
) -> Tuple[List[State], int]:
    """
    シード文節 seed から、target_bunsetsu 文節の回文候補を探索する。

    Returns
    -------
    (候補リスト, 総探索ノード数)
    """
    if target_bunsetsu < 1:
        return [], 0

    results: List[State] = []
    seen: Set[Tuple[str, str, str, int]] = set()
    init_states = initial_states(seed)
    total = len(init_states)
    explored_nodes = 0

    def dfs(st: State):
        nonlocal explored_nodes
        if cancel_callback is not None and cancel_callback():
            return

        explored_nodes += 1

        key = (st.L, st.H, st.R, len(st.parts))
        if key in seen:
            return
        seen.add(key)

        if len(st.parts) == target_bunsetsu:
            if st.is_final() and final_ok(st.parts):
                results.append(st)
                if result_callback is not None:
                    result_callback(st)
            return

        if st.L:
            left_neighbor = st.parts[0] if st.parts else None
            for w in db:
                if cancel_callback is not None and cancel_callback():
                    return
                if (not allow_duplicate_parts) and (w in st.parts):
                    continue
                if not can_connect(left_neighbor, w, "left"):
                    continue
                nxt = extend_left(st, w)
                if nxt is not None:
                    dfs(nxt)

        elif st.R:
            right_neighbor = st.parts[-1] if st.parts else None
            for w in db:
                if cancel_callback is not None and cancel_callback():
                    return
                if (not allow_duplicate_parts) and (w in st.parts):
                    continue
                if not can_connect(right_neighbor, w, "right"):
                    continue
                nxt = extend_right(st, w)
                if nxt is not None:
                    dfs(nxt)

    for idx, st0 in enumerate(init_states, start=1):
        dfs(st0)
        if progress_callback is not None:
            progress_callback(idx, total)

    uniq: List[State] = []
    emitted: Set[Tuple[str, str]] = set()
    for st in results:
        k = (st.text(), st.reading())
        if k not in emitted:
            emitted.add(k)
            uniq.append(st)

    return uniq, explored_nodes


# =============================================================================
# 補助関数群
# =============================================================================

def is_palindrome_reading(reading: str) -> bool:
    return reading == reading[::-1]



def find_seed(db: Sequence[Bunsetsu], seed_text: Optional[str] = None, seed_reading: Optional[str] = None) -> Bunsetsu:
    for bun in db:
        if seed_text is not None and seed_reading is not None:
            if bun.text == seed_text and bun.reading == seed_reading:
                return bun
        elif seed_text is not None:
            if bun.text == seed_text:
                return bun
        elif seed_reading is not None:
            if bun.reading == seed_reading:
                return bun

    raise ValueError(f"シード文節が見つかりません。seed_text={seed_text}, seed_reading={seed_reading}")



def build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="文節CSVを読み込んで回文候補を生成する")
    p.add_argument("--csv", help="文節CSVファイルのパス。未指定時は内蔵デモ辞書を使う")
    p.add_argument("--encoding", default="utf-8-sig", help="CSVの文字コード。既定: utf-8-sig")
    p.add_argument("--seed-text", help="シード文節の表記")
    p.add_argument("--seed-reading", help="シード文節の読み")
    p.add_argument("--target", type=int, default=4, help="生成する目標文節数。既定: 4")
    p.add_argument("--no-duplicate", action="store_true", help="同じ文節の再利用を禁止する")
    p.add_argument("--show-progress", action="store_true", help="進捗率を表示する")
    return p



def demo_db() -> List[Bunsetsu]:
    return [
        Bunsetsu("夜", "よる", "noun"),
        Bunsetsu("起きぬ", "おきぬ", "verb"),
        Bunsetsu("タヌキ", "たぬき", "noun"),
        Bunsetsu("おるよ", "おるよ", "aux"),
        Bunsetsu("良い", "よい", "adj"),
        Bunsetsu("滝", "たき", "noun"),
        Bunsetsu("行きたいよ", "いきたいよ", "verb"),
    ]



def make_progress_callback() -> ProgressCallback:
    last_percent = -1

    def _callback(done: int, total: int) -> None:
        nonlocal last_percent
        percent = int(done * 100 / total) if total else 100
        if percent != last_percent:
            print(f"進捗: {percent}% ({done}/{total})", file=sys.stderr)
            last_percent = percent

    return _callback


class LiveCandidatePrinter:
    """
    実行中に、最新の回文候補を標準出力へ 1 行で上書き表示するための補助クラス。

    使い方
    ------
    - 候補発見時に __call__(state) を呼ぶ
    - 最後に finalize() を呼ぶと、上書き表示用の行を改行で確定する

    実装上の工夫
    ------------
    前回表示より今回表示が短いと、ターミナル上に文字が残ることがある。
    それを防ぐため、前回表示長を記録し、余った部分を空白で上書きしている。
    """
    def __init__(self) -> None:
        self.prev_len = 0
        self.used = False

    def __call__(self, st: State) -> None:
        msg = f"生成中: {st.text()} / {st.reading()} / 自然度={compute_naturality_score(st.parts)}"
        pad = " " * max(self.prev_len - len(msg), 0)
        print("\r" + msg + pad, end="", flush=True)
        self.prev_len = len(msg)
        self.used = True

    def finalize(self) -> None:
        if self.used:
            print()  # 最後に改行して以降の通常出力と混ざらないようにする



def safe_filename(name: str) -> str:
    invalid_chars = '\\/:*?"<>|'
    out = ''.join('_' if ch in invalid_chars else ch for ch in name).strip()
    return out if out else 'seed'



def output_excel_path(input_csv_path: Optional[str], seed: Bunsetsu) -> Path:
    base_dir = Path(input_csv_path).resolve().parent if input_csv_path else Path.cwd()
    date_str = datetime.now().strftime('%Y%m%d')
    return base_dir / f"{safe_filename(seed.text)}_{date_str}.xlsx"



def _autofit_worksheet_columns(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            value = '' if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)



def write_candidates_excel(path: Path, seed: Bunsetsu, target: int, candidates: Sequence[State], elapsed_seconds: float, total_explored_nodes: int, args_dict: dict) -> None:
    """
    生成した回文候補一覧を Excel ブック (.xlsx) に保存する。
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # -------------------------------------------------------------------------
    # シート1: 一覧
    # -------------------------------------------------------------------------
    ws_list = wb.active
    ws_list.title = "一覧"

    list_headers = [
        'seed_text',
        'seed_reading',
        'target_bunsetsu',
        'candidate_no',
        'text',
        'reading',
        'is_palindrome',
        'naturality_score',
        'parts_text',
        'parts_reading',
        'parts_btype',
    ]
    ws_list.append(list_headers)

    for i, c in enumerate(candidates, start=1):
        ws_list.append([
            seed.text,
            seed.reading,
            target,
            i,
            c.text(),
            c.reading(),
            is_palindrome_reading(c.reading()),
            compute_naturality_score(c.parts),
            '|'.join(p.text for p in c.parts),
            '|'.join(p.reading for p in c.parts),
            '|'.join(p.btype for p in c.parts),
        ])

    ws_list.freeze_panes = 'A2'
    _autofit_worksheet_columns(ws_list)

    # -------------------------------------------------------------------------
    # シート2: サマリー
    # -------------------------------------------------------------------------
    ws_summary = wb.create_sheet(title="サマリー")
    ws_summary.append(['項目', '値'])
    ws_summary.append(['seed_text', seed.text])
    ws_summary.append(['seed_reading', seed.reading])
    ws_summary.append(['target_bunsetsu', target])
    ws_summary.append(['candidate_count', len(candidates)])
    ws_summary.append(['total_explored_nodes', total_explored_nodes])
    ws_summary.append(['elapsed_seconds', elapsed_seconds])
    ws_summary.append(['elapsed_hms', f'{elapsed_seconds:.3f} sec'])
    ws_summary.append(['generated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_summary.append(['output_file', str(path)])
    ws_summary.append(['', ''])
    ws_summary.append(['実行時引数', '値'])

    for key, value in args_dict.items():
        ws_summary.append([key, '' if value is None else str(value)])

    ws_summary.freeze_panes = 'A2'
    _autofit_worksheet_columns(ws_summary)

    wb.save(path)


# =============================================================================
# エントリポイント
# =============================================================================

def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_argparser().parse_args(argv)

    if args.csv:
        db = load_bunsetsu_csv(args.csv, encoding=args.encoding)
        print(f"CSVから {len(db)} 件の文節を読み込みました: {args.csv}")
    else:
        db = demo_db()
        print(f"内蔵デモ辞書 {len(db)} 件を使用します")

    if args.seed_text or args.seed_reading:
        seed = find_seed(db, seed_text=args.seed_text, seed_reading=args.seed_reading)
    else:
        try:
            seed = find_seed(db, seed_text="タヌキ")
        except ValueError:
            seed = db[0]

    print(f"シード文節: {seed.text} / {seed.reading} / {seed.btype}")
    print(f"目標文節数: {args.target}")

    progress_cb = make_progress_callback() if args.show_progress else None
    live_printer = LiveCandidatePrinter()

    start_time = time.perf_counter()
    cands, explored_nodes = generate_palindrome_candidates(
        seed=seed,
        db=db,
        target_bunsetsu=args.target,
        allow_duplicate_parts=not args.no_duplicate,
        progress_callback=progress_cb,
        result_callback=live_printer,
    )
    elapsed_seconds = time.perf_counter() - start_time

    # 上書き表示に使っていた行を確定して、以降の通常出力を見やすくする
    live_printer.finalize()

    print(f"候補数: {len(cands)}")
    # for i, c in enumerate(cands, start=1):
    #     print(f"[{i}] {c.text()}\t{c.reading()}\t回文={is_palindrome_reading(c.reading())}\t自然度={compute_naturality_score(c.parts)}")

    out_path = output_excel_path(args.csv, seed)
    write_candidates_excel(out_path, seed, args.target, cands, elapsed_seconds, explored_nodes, vars(args).copy())
    print(f"処理時間: {elapsed_seconds:.3f} 秒")
    print(f"総探索ノード数: {explored_nodes}")
    print(f"結果Excelを保存しました: {out_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
